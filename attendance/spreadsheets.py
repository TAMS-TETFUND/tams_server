from datetime import datetime
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from db.models import AttendanceSession, EventTypeChoices


def generate_attendance_records_sheet(
    attendance_session: AttendanceSession, attendance_records_qs
) -> Workbook:
    template_path = os.path.join(
        "static", "spreadsheets", "attendance_template.xlsx"
    )
    wb = load_workbook(template_path)
    ws = wb.active
    body_font = Font(size="9")

    # course field
    ws["D4"] = "{} - {}".format(
        attendance_session.course.code, attendance_session.course.title
    )
    ws["D4"].font = body_font
    ws["D4"].alignment = Alignment(wrap_text=True, horizontal="center")

    # event type field
    ws["D5"] = EventTypeChoices(attendance_session.event_type).label
    # staff name field
    ws["D6"] = "{}. {}".format(
        attendance_session.initiator.first_name[0],
        attendance_session.initiator.last_name,
    )

    # lecture date/time
    ws["W4"] = attendance_session.start_time.strftime("%d-%b-%Y %H:%M")
    ws["W4"].font = body_font

    # academic session
    ws["W5"] = "{}".format(attendance_session.session.session)

    attendance_columns = [1, 3, 14, 19, 25, 29]

    row = 9
    col = 1

    formatted_att_list = []
    for idx, dict_el in enumerate(attendance_records_qs, 1):
        formatted_att_list.append(
            {
                "S/N": idx,
                "Name": f'{dict_el["student__last_name"].capitalize()} {dict_el["student__first_name"].capitalize()}',
                "Reg. Number": dict_el["student__reg_number"],
                "Department": dict_el["student__department__name"],
                "Sign In": f'{datetime.strftime(dict_el["check_in_by"], "%H:%M")}',
            }
        )

    for record in formatted_att_list:
        for idx, value in enumerate(record.values()):
            _ = ws.cell(row=row, column=attendance_columns[idx], value=value)
            _.alignment = Alignment(horizontal="center")
            _.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            _.font = Font("Calibri", 9)

    # merge after adding border to maintain border across merged cells
    for r_idx in range(row, (row + len(attendance_records_qs) + 15)):
        current_col = attendance_columns[0]
        for col_idx in attendance_columns[1:]:
            cell_start = get_column_letter(current_col) + str(r_idx)
            cell_stop = get_column_letter(col_idx - 1) + str(r_idx)
            _range = ":".join([cell_start, cell_stop])
            ws.merge_cells(_range)
            current_col = col_idx

    return wb
